import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Products"

# 1. 写入表头
worksheet.append(["Product", "Quantity", "Price"])

# 2. 数据资料
products_data = [
    ["盒装纸巾 A", 10, 124],
    ["盒装纸巾 B", 10, 124],
]

for row in products_data:
    worksheet.append(row)

# 3. 动态获取最后一行
last_row = worksheet.max_row

# 4. 写入总计
total_row = last_row + 2
worksheet.cell(row=total_row, column=1, value="总共消费").font = Font(name="Arial", size=14, bold=True)

# 【核心修正】：使用 SUMPRODUCT 代替 SUM*SUM
sumproduct_formula = f"=SUMPRODUCT(B2:B{last_row}, C2:C{last_row})"
worksheet.cell(row=total_row, column=2, value=sumproduct_formula)

# 6. 添加 styles
# 定义边框样式（细黑线）
thin_side = Side(border_style="thin", color="000000")
borders = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# 定义表头底色（浅蓝色）
header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

# 循环处理每一行、每一列，应用样式
# 从第 1 行到总计行，从第 1 列到第 3 列
for r in range(1, total_row + 1):
    for c in range(1, 4):
        cell = worksheet.cell(row=r, column=c)
        
        # 1. 给所有有数据的单元格加边框
        if cell.value is not None or r == total_row:
            cell.border = borders
            
        # 2. 居中对齐
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 3. 如果是第一行（表头），加粗并上色
        if r == 1:
            cell.font = Font(bold=True)
            cell.fill = header_fill

# 7. 自动调整列宽（让 A 栏不再遮挡文字）
worksheet.column_dimensions['A'].width = 20
worksheet.column_dimensions['B'].width = 12
worksheet.column_dimensions['C'].width = 12
 
# 5. 保存
# workbook.save("product_list.xlsx")
# print("Bug 已修复！现在使用的是 SUMPRODUCT 逻辑。")

# 6. 把第五点变成函数
def generate_excel():
    workbook.save("product_list.xlsx")
    print("Bug 已修复！现在使用的是 SUMPRODUCT 逻辑。")